"""Column statistics for structured formats."""

from __future__ import annotations

import os
from typing import Any

from rich.console import Console
from rich.table import Table

console = Console()

UNIQUE_CAP = 10_000


def _fmt_num(val: Any) -> str:
    """Format a number with thousands separators."""
    if val is None:
        return "—"
    if isinstance(val, float):
        return f"{val:,.2f}"
    if isinstance(val, int):
        return f"{val:,}"
    return str(val)


def _supports_mean(type_str: str) -> bool:
    """Check if a type supports mean computation."""
    t = type_str.upper()
    for kw in ("INT", "FLOAT", "DOUBLE", "DECIMAL", "NUM"):
        if kw in t:
            return True
    return False


def _print_stats_table(
    filename: str,
    total_rows: int,
    total_cols: int,
    col_stats: list[dict],
    file_size: int | None = None,
    fmt: str | None = None,
    compression: str | None = None,
):
    """Render the stats Rich table."""
    table = Table(
        title=f"Stats  {filename}  ({_fmt_num(total_rows)} rows · {total_cols} columns)",
        show_header=True,
        header_style="bold cyan",
    )
    table.add_column("Column")
    table.add_column("Type")
    table.add_column("Non-Null", justify="right")
    table.add_column("Null", justify="right")
    table.add_column("Min")
    table.add_column("Max")
    table.add_column("Mean")
    table.add_column("Unique", justify="right")

    for cs in col_stats:
        null_count = cs.get("null_count", 0)
        non_null = cs.get("non_null", total_rows - null_count)
        null_str = _fmt_num(null_count)
        if null_count > 0 and total_rows > 0:
            pct = null_count / total_rows * 100
            null_str += f" ({pct:.1f}%)"

        mean_val = cs.get("mean")
        mean_str = _fmt_num(mean_val) if mean_val is not None else "—"

        unique = cs.get("unique")
        if unique is None:
            unique_str = "—"
        elif unique > UNIQUE_CAP:
            unique_str = f"{UNIQUE_CAP:,}+"
        else:
            unique_str = _fmt_num(unique)

        table.add_row(
            cs.get("name", ""),
            cs.get("type", ""),
            _fmt_num(non_null),
            null_str,
            str(cs.get("min", "—")) if cs.get("min") is not None else "—",
            str(cs.get("max", "—")) if cs.get("max") is not None else "—",
            mean_str,
            unique_str,
        )

    console.print(table)

    # Footer
    parts = []
    if file_size is not None:
        if file_size > 1_000_000_000:
            parts.append(f"{file_size / 1_000_000_000:.1f} GB")
        elif file_size > 1_000_000:
            parts.append(f"{file_size / 1_000_000:.1f} MB")
        elif file_size > 1_000:
            parts.append(f"{file_size / 1_000:.1f} KB")
        else:
            parts.append(f"{file_size} B")
    if fmt:
        parts.append(fmt)
    if compression:
        parts.append(f"compression: {compression}")
    if parts:
        console.print(f"[dim]{' · '.join(parts)}[/dim]")


def stats_parquet(path: str, columns: list[str] | None = None, s3_endpoint: str | None = None):
    """Compute stats from Parquet metadata (no data read)."""
    import pyarrow.parquet as pq

    if "://" in path:
        import fsspec
        so = {}
        if s3_endpoint and path.startswith("s3://"):
            so["client_kwargs"] = {"endpoint_url": s3_endpoint}
        fs, fpath = fsspec.core.url_to_fs(path, **so)
        pf = pq.ParquetFile(fs.open(fpath, "rb"))
        file_size = fs.size(fpath)
    else:
        pf = pq.ParquetFile(path)
        file_size = os.path.getsize(path)

    meta = pf.metadata
    schema = pf.schema_arrow
    total_rows = meta.num_rows

    # Build column index mapping
    col_names = [schema.field(i).name for i in range(len(schema))]
    col_types = [str(schema.field(i).type) for i in range(len(schema))]

    if columns:
        indices = [i for i, n in enumerate(col_names) if n in columns]
    else:
        indices = list(range(len(col_names)))

    # Get compression from first column of first row group
    compression = None
    if meta.num_row_groups > 0 and meta.num_columns > 0:
        compression = meta.row_group(0).column(0).compression

    col_stats_list = []
    for col_idx in indices:
        stats_agg: list = []
        null_total = 0
        for rg in range(meta.num_row_groups):
            rg_meta = meta.row_group(rg)
            col_meta = rg_meta.column(col_idx)
            if col_meta.statistics:
                stats_agg.append(col_meta.statistics)
                null_total += col_meta.statistics.null_count

        cs: dict[str, Any] = {
            "name": col_names[col_idx],
            "type": col_types[col_idx].upper(),
            "null_count": null_total,
            "non_null": total_rows - null_total,
        }

        if stats_agg:
            mins = [s.min for s in stats_agg if s.has_min_max]
            maxs = [s.max for s in stats_agg if s.has_min_max]
            if mins:
                cs["min"] = min(mins)
            if maxs:
                cs["max"] = max(maxs)

            # Distinct count from stats if available
            distinct = [s.num_values for s in stats_agg]
            # num_values is non-null count per RG, not unique — skip for unique

        # Mean only for numeric types
        if _supports_mean(cs["type"]):
            # Can't compute mean from metadata alone without reading data
            cs["mean"] = None

        col_stats_list.append(cs)

    _print_stats_table(
        os.path.basename(path),
        total_rows,
        len(indices),
        col_stats_list,
        file_size=file_size,
        fmt="parquet",
        compression=compression,
    )


def stats_streaming(path: str, fmt: str, columns: list[str] | None = None, s3_endpoint: str | None = None, file_obj=None):
    """Compute stats by streaming through the file (CSV, JSONL, Avro, ORC)."""
    import numbers

    file_size = None
    if "://" not in path and os.path.exists(path):
        file_size = os.path.getsize(path)

    # Accumulators per column
    accum: dict[str, dict[str, Any]] = {}
    total_rows = 0

    def _update(name: str, value: Any):
        if name not in accum:
            accum[name] = {
                "min": None, "max": None, "sum": 0.0, "num_count": 0,
                "null_count": 0, "non_null": 0, "uniques": set(), "unique_capped": False,
                "type": None,
            }
        a = accum[name]
        if value is None or (isinstance(value, float) and value != value):
            a["null_count"] += 1
            return
        a["non_null"] += 1

        # Type detection
        if a["type"] is None:
            if isinstance(value, bool):
                a["type"] = "BOOL"
            elif isinstance(value, int):
                a["type"] = "INT64"
            elif isinstance(value, float):
                a["type"] = "FLOAT64"
            elif isinstance(value, str):
                a["type"] = "STRING"
            else:
                a["type"] = type(value).__name__.upper()

        # Min/Max
        try:
            if a["min"] is None or value < a["min"]:
                a["min"] = value
            if a["max"] is None or value > a["max"]:
                a["max"] = value
        except TypeError:
            pass

        # Mean accumulator for numeric
        if isinstance(value, numbers.Number) and not isinstance(value, bool):
            a["sum"] += float(value)
            a["num_count"] += 1

        # Unique tracking
        if not a["unique_capped"]:
            try:
                a["uniques"].add(value)
                if len(a["uniques"]) > UNIQUE_CAP:
                    a["unique_capped"] = True
            except TypeError:
                a["unique_capped"] = True

    def _iter_rows():
        nonlocal file_obj
        from mcat.structured import _open_file, _check_extra

        if fmt == "csv" or fmt == "tsv":
            import csv as csv_mod
            delimiter = "\t" if fmt == "tsv" else ","
            if file_obj:
                import io
                f = io.TextIOWrapper(file_obj, encoding="utf-8")
            else:
                f = _open_file(path, "r", s3_endpoint=s3_endpoint)
            reader = csv_mod.DictReader(f, delimiter=delimiter)
            for row in reader:
                yield row
            if not file_obj:
                f.close()
        elif fmt == "jsonl":
            import json
            if file_obj:
                import io
                f = io.TextIOWrapper(file_obj, encoding="utf-8")
            else:
                f = _open_file(path, "r", s3_endpoint=s3_endpoint)
            for line in f:
                line = line.strip()
                if line:
                    try:
                        yield json.loads(line)
                    except Exception:
                        pass
            if not file_obj:
                f.close()
        elif fmt == "avro":
            _check_extra("Avro", "fastavro", "avro")
            import fastavro
            if file_obj:
                f = file_obj
            else:
                f = _open_file(path, "rb", s3_endpoint=s3_endpoint)
            reader = fastavro.reader(f)
            for record in reader:
                yield record
            if not file_obj:
                f.close()
        elif fmt == "orc":
            import pyarrow.orc as orc
            if file_obj:
                f = file_obj
            else:
                f = _open_file(path, "rb", s3_endpoint=s3_endpoint)
            reader = orc.ORCFile(f)
            table = reader.read()
            rows_dict = table.to_pydict()
            if rows_dict:
                keys = list(rows_dict.keys())
                n = len(rows_dict[keys[0]])
                for i in range(n):
                    yield {k: rows_dict[k][i] for k in keys}
            if not file_obj:
                f.close()
        elif fmt == "excel":
            is_xls = path.split("?")[0].split("#")[0].lower().endswith(".xls")
            if is_xls:
                _check_extra("Excel (.xls)", "xlrd", "excel")
                import xlrd
                f = file_obj if file_obj else _open_file(path, "rb", s3_endpoint=s3_endpoint)
                data = f.read()
                if not file_obj:
                    f.close()
                wb = xlrd.open_workbook(file_contents=data)
                sheet = wb.sheet_by_index(0)
                headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
                for r in range(1, sheet.nrows):
                    yield {headers[c]: sheet.cell_value(r, c) for c in range(sheet.ncols)}
            else:
                _check_extra("Excel (.xlsx)", "openpyxl", "excel")
                import openpyxl
                f = file_obj if file_obj else _open_file(path, "rb", s3_endpoint=s3_endpoint)
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                row_iter = ws.iter_rows(values_only=True)
                headers = [str(c) if c is not None else "" for c in next(row_iter)]
                for values in row_iter:
                    yield {headers[i]: values[i] for i in range(len(headers))}
                wb.close()
                if not file_obj:
                    f.close()
        elif fmt in ("feather", "arrow"):
            import pyarrow
            import pyarrow.feather
            import pyarrow.ipc
            f = file_obj if file_obj else _open_file(path, "rb", s3_endpoint=s3_endpoint)
            clean = path.split("?")[0].split("#")[0].lower()
            if clean.endswith(".arrow") or fmt == "arrow":
                table = pyarrow.ipc.open_stream(f).read_all()
            else:
                table = pyarrow.feather.read_table(f)
            rows_dict = table.to_pydict()
            if rows_dict:
                keys = list(rows_dict.keys())
                n = len(rows_dict[keys[0]])
                for i in range(n):
                    yield {k: rows_dict[k][i] for k in keys}
            if not file_obj:
                f.close()
        elif fmt == "json":
            import json
            if file_obj:
                import io
                f = io.TextIOWrapper(file_obj, encoding="utf-8")
            else:
                f = _open_file(path, "r", s3_endpoint=s3_endpoint)
            data = json.load(f)
            if isinstance(data, dict):
                data = [data]
            for row in data:
                yield row
            if not file_obj:
                f.close()

    for row in _iter_rows():
        total_rows += 1
        for k, v in row.items():
            if columns and k not in columns:
                continue
            _update(k, v)

    # Build stats list
    if columns:
        ordered_keys = [c for c in columns if c in accum]
    else:
        ordered_keys = list(accum.keys())

    col_stats_list = []
    for name in ordered_keys:
        a = accum[name]
        cs: dict[str, Any] = {
            "name": name,
            "type": a["type"] or "UNKNOWN",
            "null_count": a["null_count"],
            "non_null": a["non_null"],
            "min": a["min"],
            "max": a["max"],
        }
        if _supports_mean(cs["type"]) and a["num_count"] > 0:
            cs["mean"] = a["sum"] / a["num_count"]
        if a["unique_capped"]:
            cs["unique"] = UNIQUE_CAP + 1
        else:
            cs["unique"] = len(a["uniques"]) if a["uniques"] else None

        col_stats_list.append(cs)

    _print_stats_table(
        os.path.basename(path),
        total_rows,
        len(ordered_keys),
        col_stats_list,
        file_size=file_size,
        fmt=fmt,
    )


def handle_stats(path: str, fmt: str, columns: list[str] | None = None, s3_endpoint: str | None = None, file_obj=None):
    """Main entry point for --stats."""
    if fmt == "parquet" and file_obj is None:
        stats_parquet(path, columns=columns, s3_endpoint=s3_endpoint)
    else:
        stats_streaming(path, fmt, columns=columns, s3_endpoint=s3_endpoint, file_obj=file_obj)
